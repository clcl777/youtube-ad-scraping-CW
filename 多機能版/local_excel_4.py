from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import urllib.request
import time
import re
import openpyxl
import datetime
import pyautogui
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

#Excelファイル名決定
k = 1
while True:
    if os.path.isfile('Excel//ad_4' + str(k) + '.xlsx'):
        k = k + 1
    else:
        file_name = 'Excel//ad_4' + str(k) + '.xlsx'
        #file_name2 = 'Excel//ad' + str(k) + '_割合.xlsx'
        #フォルダ作成
        try:
            os.mkdir('image//ad' + str(k))
            os.mkdir('image//ad' + str(k) + '//元の広告動画')
            os.mkdir('image//ad' + str(k) + '//LP')
            os.mkdir('image//ad' + str(k) + '//先の広告動画')
            os.mkdir('image//ad' + str(k) + '//error')
        except:
            print('フォルダ作成済み')
        break


pyautogui.FAILSAFE = False

#連続カウント
ct_error = 0
#広告カウント
ct_ad = 1
#UNIX時間取得
ut = time.time()
dt_now = datetime.datetime.now()
today = dt_now.strftime('%Y年%m月%d日')

initial_url = 'https://www.youtube.com/watch?v=cjGUgsTiqm8'
#リスト変更
# url_lists = ['https://www.youtube.com/watch?v=riPjiA3aZWU','https://www.youtube.com/watch?v=M3NIsBA72I4&t=3s','https://www.youtube.com/watch?v=w-o73uuIREU']

#initial_url = 'https://www.youtube.com/watch?v=OUWieJFqs9c / QDDY NNHB 2RYV'
options = Options()
driver = webdriver.Chrome(executable_path="C:\\Users\\tisk0\\PycharmProjects\\chromedriver_win32\\chromedriver.exe",options=options)
driver.maximize_window()
driver.implicitly_wait(10)
driver.get(initial_url)

wb = openpyxl.Workbook()
ws = wb['Sheet']
ws.cell(1, 1, value='広告URL')
ws.cell(1, 2, value='再生回数')
ws.cell(1, 3, value='投稿日')
ws.cell(1, 4, value='限定公開')
ws.cell(1, 5, value='動画タイトル')
ws.cell(1, 6, value='投稿者名')
ws.cell(1, 7, value='動画長さ')
ws.cell(1, 8, value='クリック先URL')
ws.cell(1, 9, value=today)
ws.cell(1, 10, value=ut)

#wb2 = openpyxl.Workbook()
#ws2 = wb2['Sheet']
#ws2.cell(1, 2, value='広告割合')
#ws2.cell(2,2).number_format = '0.0'
#ws2.cell(2, 2, value='=ROUND(COUNTIF(A:A,"広告あり")/COUNTA(A:A)*100,1)&"%"')
l = 1
# 詳細統計表示
time.sleep(3)
webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
time.sleep(3)
video_element = driver.find_element_by_css_selector('.video-stream.html5-main-video')
webdriver.ActionChains(driver).context_click(video_element).perform()
# 786 838
# pyautogui.click(786, 838)
time.sleep(1)
syosaitouke_elements = driver.find_elements_by_css_selector('.ytp-menuitem-label')
for syosaitouke_element in syosaitouke_elements:
    if syosaitouke_element.get_attribute('textContent') == '詳細統計情報':
        syosaitouke_element.click()
while True:
    #再生ボタン押す
    try:
        time.sleep(3)
        #pyautogui.press('space')
        webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
        time.sleep(3)
        #広告かどうか
        #ad_marker = driver.find_elements_by_css_selector('#button\:i')
        #pyautogui.press('space')
        webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
        ad_marker = driver.find_elements_by_class_name('ytp-ad-player-overlay-instream-info')
        if len(ad_marker) > 0:
            print('広告あり')
            #ws2.cell(l, 1, value='広告あり')
            #スクショ1 元の広告動画
            driver.save_screenshot('image//ad' + str(k) + '//元の広告動画//' + str(ct_ad) + '.png')
            #広告URL取得
            ad_url_element = driver.find_element_by_css_selector(
                '#movie_player > div.html5-video-info-panel > div > div:nth-child(1) > span')
            ad_url = ad_url_element.get_attribute('textContent')
            ad_url = 'https://www.youtube.com/watch?v=' + ad_url
            ws.cell(ct_ad + 1, 1, value=ad_url)

            '''
            #650 516
            pyautogui.rightClick(650, 516)
            time.sleep(2)
            pyautogui.click(802, 671)
            ad_url_elements = driver.find_elements_by_css_selector('#movie_player > div.html5-video-info-panel > div > div:nth-child(1) > span')
            if len(ad_url_elements):
                ad_url = ad_url_elements[0].get_attribute('textContent')
                ad_url = 'https://www.youtube.com/watch?v=' + ad_url
            else:
                pyautogui.rightClick(650, 516)
                #786 838
                pyautogui.click(786, 838)
                ad_url_element = driver.find_element_by_css_selector('#movie_player > div.html5-video-info-panel > div > div:nth-child(1) > span')
                ad_url = ad_url_element.get_attribute('textContent')
                ad_url = 'https://www.youtube.com/watch?v=' + ad_url
            '''
            #ad_url = 'https://www.youtube.com/watch?v=Pc6mW5XW3Ug / 5YNB 62M7 ASV1'
            #driver.get(ad_url)
            time.sleep(1)

            # クリック先URL
            lp_element = driver.find_element_by_class_name('ytp-ad-button-text')
            #lp_element = driver.find_element_by_css_selector('.ytp-ad-button.ytp-ad-visit-advertiser-button.ytp-ad-button-link')
            #lp_element.click()
            #LP移動
            driver.execute_script("arguments[0].click();", lp_element)
            driver.switch_to.window(driver.window_handles[1])
            time.sleep(3)
            # スクショ2　LP
            driver.save_screenshot('image//ad' + str(k) + '//LP//' + str(ct_ad) + '.png')
            lp_url = driver.current_url
            ws.cell(ct_ad + 1, 8, value=lp_url)
            # lp_element = driver.find_element_by_css_selector('#visit-advertiser\:p > span.ytp-ad-button-text')
            # lp = lp_element.get_attribute('textContent')
            print(lp_url)
            #driver.close()
            #LPからメインへ移動
            #driver.switch_to.window(driver.window_handles[0])
            driver.get(ad_url)

            #スクショ3 先の広告動画
            driver.save_screenshot('image//ad' + str(k) + '//先の広告動画//' + str(ct_ad) + '-3.png')
            #再生回数
            views_element = driver.find_element_by_css_selector('#count > ytd-video-view-count-renderer > span.view-count.style-scope.ytd-video-view-count-renderer')
            views = views_element.get_attribute('textContent')
            views = views[:-4]
            views = views.replace(",", "")
            views = int(views)
            print(views)
            ws.cell(ct_ad + 1, 2, value=views)

            #投稿日
            date_element = driver.find_element_by_css_selector('#info-strings > yt-formatted-string')
            date = date_element.get_attribute('textContent')
            print(date)
            ws.cell(ct_ad + 1, 3, value=date)

            #限定公開
            limited_elements = driver.find_elements_by_css_selector('#container > ytd-badge-supported-renderer:nth-child(5) > div')
            if len(limited_elements) > 0:
                print('限定あり')
                ws.cell(ct_ad + 1, 4, value='〇')
            else:
                print('限定なし')
                ws.cell(ct_ad + 1, 4, value='×')

            #動画タイトル
            title_element = driver.find_element_by_css_selector('#container > h1 > yt-formatted-string')
            title = title_element.get_attribute('textContent')
            print(title)
            ws.cell(ct_ad + 1, 5, value=title)

            #投稿者名
            uploader_element = driver.find_element_by_css_selector('#text > a')
            uploader = uploader_element.get_attribute('textContent')
            print(uploader)
            ws.cell(ct_ad + 1, 6, value=uploader)

            #動画長さ
            length_element = driver.find_element_by_css_selector('#movie_player > div.ytp-chrome-bottom > div.ytp-chrome-controls > div.ytp-left-controls > div.ytp-time-display.notranslate > span.ytp-time-duration')
            length = length_element.get_attribute('textContent')
            print(length)
            ws.cell(ct_ad + 1, 7, value=length)

            ct_error = 0

            #次の動画へ
            wb.save(file_name)
            ct_ad = ct_ad + 1
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
            #動画と動画の間隔
            time.sleep(180)
            #pyautogui.hotkey('shift', 'n')
            webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()
            #driver.get(url_lists[random.randrange(3)])


        else:
            print('広告なし')
            #ws2.cell(l, 1, value='広告なし')
            next_video_element = driver.find_element_by_css_selector('#thumbnail')
            next_video_url = next_video_element.get_attribute('href')
            # print(next_video_url)
            #driver.get(next_video_url)  # エラー起きたとき指定のURLを開く
            time.sleep(10)
            #pyautogui.hotkey('shift', 'n')#次の動画
            webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()
            time.sleep(1)

    except:
        print('エラー発生')
        #ws2.cell(l, 1, value='エラー発生')
        ct_error = ct_error + 1
        #エラースクショ
        i = 1
        while True:
            if os.path.isfile('image//ad' + str(k) + '//error//' + str(i) + '.png'):
                i = i + 1
            else:
                driver.save_screenshot('image//ad' + str(k) + '//error//' + str(i) + '.png')
                #wb.save('ad' + str(k) + '.xlsx')
                break
        time.sleep(10)
        if ct_error > 5:
            driver.get(initial_url)
            time.sleep(5)
            # 詳細統計表示
            video_element = driver.find_element_by_css_selector('.video-stream.html5-main-video')
            webdriver.ActionChains(driver).context_click(video_element).perform()
            # 786 838
            # pyautogui.click(786, 838)
            time.sleep(1)
            syosaitouke_elements = driver.find_elements_by_css_selector('.ytp-menuitem-label')
            for syosaitouke_element in syosaitouke_elements:
                if syosaitouke_element.get_attribute('textContent') == '詳細統計情報':
                    syosaitouke_element.click()

        else:
            #pyautogui.hotkey('shift', 'n')  # 次の動画
            webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()
    #wb2.save(file_name2)
    l = l + 1