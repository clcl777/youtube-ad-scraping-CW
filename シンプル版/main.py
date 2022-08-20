import logging
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import sys


#Excelファイル名決定
k = 1

#連続カウント
ct_error = 0

#入力読み込み
wb_input = openpyxl.load_workbook('入力.xlsx')
ws_input = wb_input.worksheets[0]
initial_url_cell = ws_input['A'][1:]
initial_url_list = []
for cell in initial_url_cell:
    if not (cell.value == '' or cell.value is None):
        initial_url_list.append(cell.value)
#出力読み込み
#ない場合作成
if os.path.isfile('出力.xlsx'):
    wb_output = openpyxl.load_workbook('出力.xlsx')
    ws_output = wb_output.worksheets[0]
    num_row = len(ws_output['A'])
    # 広告カウント
    ct_ad = num_row
else:
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.worksheets[0]
    ct_ad = 1

options = Options()
options.add_argument('--headless')
options.add_argument("--mute-audio")
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver.maximize_window()
driver.implicitly_wait(10)

l = 1
'''
# 詳細統計表示
time.sleep(3)
webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
time.sleep(3)
video_element = driver.find_element_by_css_selector('.video-stream.html5-main-video')
webdriver.ActionChains(driver).context_click(video_element).perform()
time.sleep(1)
syosaitouke_elements = driver.find_elements_by_css_selector('.ytp-menuitem-label')
for syosaitouke_element in syosaitouke_elements:
    if syosaitouke_element.get_attribute('textContent') == '詳細統計情報':
        syosaitouke_element.click()
'''
#広告数
ad_num = 1
while True:
    for initial_url in initial_url_list:
        driver.get(initial_url)
        time.sleep(2)
        webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
        time.sleep(2)
        video_element = driver.find_element_by_css_selector('.video-stream.html5-main-video')
        webdriver.ActionChains(driver).context_click(video_element).perform()
        time.sleep(1)
        syosaitouke_elements = driver.find_elements_by_css_selector('.ytp-menuitem-label')
        for syosaitouke_element in syosaitouke_elements:
            if syosaitouke_element.get_attribute('textContent') == '詳細統計情報':
                syosaitouke_element.click()
        #再生ボタン押す
        try:
            time.sleep(3)
            #pyautogui.press('space')
            webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
            time.sleep(3)
            #広告かどうか
            #ad_marker = driver.find_elements_by_css_selector('#button\:i')
            #pyautogui.press('space')
            webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
            ad_marker = driver.find_elements_by_class_name('ytp-ad-player-overlay-instream-info')
            if len(ad_marker) > 0:
                print('広告あり' + str(ad_num) + '件目')
                ad_num = ad_num + 1
                #ws2.cell(l, 1, value='広告あり')
                #広告URL取得
                ad_url_element = driver.find_element_by_css_selector(
                    '#movie_player > div.html5-video-info-panel > div > div:nth-child(1) > span')
                ad_url = ad_url_element.get_attribute('textContent')
                ad_url = 'https://www.youtube.com/watch?v=' + ad_url
                ad_url = ad_url.replace(' ', '')
                ws_output.cell(ct_ad + 1, 1, value=ad_url)

                time.sleep(1)


                # クリック先URL
                lp_element = driver.find_element_by_class_name('ytp-ad-button-text')

                #LP移動
                driver.execute_script("arguments[0].click();", lp_element)
                driver.switch_to.window(driver.window_handles[1])
                time.sleep(3)
                lp_url = driver.current_url
                ws_output.cell(ct_ad + 1, 10, value=lp_url)
                #print(lp_url)
                #driver.close()
                #LPからメインへ移動
                #driver.switch_to.window(driver.window_handles[0])

                driver.get(ad_url)

                #再生回数
                views_element = driver.find_element_by_css_selector('#count > ytd-video-view-count-renderer > span.view-count.style-scope.ytd-video-view-count-renderer')
                views = views_element.get_attribute('textContent')
                views = views[:-4]
                views = views.replace(",", "")
                views = int(views)
                #print(views)
                ws_output.cell(ct_ad + 1, 2, value=views)

                #投稿日
                date_element = driver.find_element_by_css_selector('#info-strings > yt-formatted-string')
                date = date_element.get_attribute('textContent')
                #print(date)
                ws_output.cell(ct_ad + 1, 4, value=date)

                #限定公開
                limited_elements = driver.find_elements_by_css_selector('#container > ytd-badge-supported-renderer:nth-child(5) > div')
                if len(limited_elements) > 0:
                    #print('限定あり')
                    ws_output.cell(ct_ad + 1, 5, value='〇')
                else:
                    #print('限定なし')
                    ws_output.cell(ct_ad + 1, 5, value='×')

                #動画タイトル
                title_element = driver.find_element_by_css_selector('#container > h1 > yt-formatted-string')
                title = title_element.get_attribute('textContent')
                #print(title)
                ws_output.cell(ct_ad + 1, 3, value=title)

                #投稿者名
                uploader_element = driver.find_element_by_css_selector('#text > a')
                uploader = uploader_element.get_attribute('textContent')
                #print(uploader)
                ws_output.cell(ct_ad + 1, 6, value=uploader)

                #動画長さ
                #length_element = driver.find_element_by_css_selector('#movie_player > div.ytp-chrome-bottom > div.ytp-chrome-controls > div.ytp-left-controls > div.ytp-time-display.notranslate > span.ytp-time-duration')
                length_element = driver.find_element_by_css_selector('#movie_player > div.ytp-chrome-bottom > div.ytp-chrome-controls > div.ytp-left-controls > div.ytp-time-display.notranslate > span:nth-child(2) > span.ytp-time-duration')
                length = length_element.get_attribute('textContent')
                #print(length)
                length_list = length.split(':')
                length = 60 * int(length_list[0]) + int(length_list[1])
                #print(length)
                ws_output.cell(ct_ad + 1, 7, value=length)

                # 評価数
                reviews_elements = driver.find_elements_by_css_selector('#text')
                reviews_list = []
                for element in reviews_elements:
                    element_text = element.get_attribute('textContent')
                    if '万' in element_text:
                        element_text = element_text.replace('万', '')
                        element_text = float(element_text) * 10000
                        element_text = int(element_text)
                        element_text = str(element_text)
                    if element_text.isdigit():
                        reviews_list.append(element_text)
                ws_output.cell(ct_ad + 1, 8, value=int(reviews_list[0]))
                ws_output.cell(ct_ad + 1, 9, value=int(reviews_list[1]))

                ct_error = 0

                #次の動画へ
                wb_output.save('出力.xlsx')
                ct_ad = ct_ad + 1
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                webdriver.ActionChains(driver).key_down(Keys.SPACE).perform()
                #動画と動画の間隔
                time.sleep(60)
                webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()

            else:
                print('広告なし')
                next_video_element = driver.find_element_by_css_selector('#thumbnail')
                next_video_url = next_video_element.get_attribute('href')
                # print(next_video_url)
                #driver.get(next_video_url)  # エラー起きたとき指定のURLを開く
                time.sleep(10)
                #pyautogui.hotkey('shift', 'n')#次の動画
                #webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()
                time.sleep(1)

        except:
            #print('エラー発生')
            print('')
            driver.close()
            driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
            driver.maximize_window()
            driver.implicitly_wait(10)
            #logging.exception("What is doing when exception happens.")
            #ws2.cell(l, 1, value='エラー発生')
            ct_error = ct_error + 1
            #エラースクショ
            i = 1
            time.sleep(10)
            if ct_error > 5:
                driver.get(initial_url)
                time.sleep(5)
                '''
                # 詳細統計表示
                video_element = driver.find_element_by_css_selector('.video-stream.html5-main-video')
                webdriver.ActionChains(driver).context_click(video_element).perform()
                # 786 838
                # pyautogui.click(786, 838)
                time.sleep(1)
                syosaitouke_elements = driver.find_elements_by_css_selector('.ytp-menuitem-label')
                for syosaitouke_element in syosaitouke_elements:
                    if syosaitouke_element.get_attribute('textContent') == '詳細統計情報':
                        syosaitouke_element.click()
                '''

            else:
                print('')
                #pyautogui.hotkey('shift', 'n')  # 次の動画
                #webdriver.ActionChains(driver).key_down(Keys.SHIFT).send_keys('n').perform()
        #wb2.save(file_name2)
        l = l + 1